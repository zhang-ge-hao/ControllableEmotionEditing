from __future__ import annotations

import os
import k_diffusion as K
from omegaconf import OmegaConf
from tqdm import tqdm
from PIL import Image, ImageDraw, ImageFont
import random
from openpyxl import Workbook
from openpyxl.drawing.image import Image

random.seed(1999)

input_root_dir = "data/0.ilsvrc"
addition_root_dir = "data/1.addition"
annotation_root_dir = "data/A.annatation"

output_dirs_and_techniques = [
    ("data/5.ultra",    "UltraEdit"),
    ("data/7.brush",    "MagicBrush"),
    ("data/6.ip2p",     "IP2P"),
    ("data/4.edited",   "Ours"),
]

select_per_emotion = 12

emotion_label_list = [
    "amusement", "awe", "contentment", "excitement", "anger", "disgust", "fear", "sadness"
]

def get_annotation(techniques, texts, image_path1_list, image_path2_list, output_path):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 10  # First column (Text) width
    ws.column_dimensions['B'].width = 10  # Second column (Image 1) width
    ws.column_dimensions['C'].width = 70  # Third column (Image 2) width
    ws.column_dimensions['D'].width = 70  # Fourth column (Image 3) width
    ws.append(["Technique", "Text", "Image 1", "Image 2"])
    for i, (technique, text, image_path1, image_path2) in enumerate(zip(techniques, texts, image_path1_list, image_path2_list)):
        ws.row_dimensions[i+2].height = 400
        ws.cell(row=i+2, column=1, value=technique)
        ws.cell(row=i+2, column=2, value=text)
        ws.add_image(Image(image_path1), f"C{i+2}")
        ws.add_image(Image(image_path2), f"D{i+2}")
    wb.save(output_path)

emotion_2_file_names = {}
for emotion in emotion_label_list:
    file_names = os.listdir(os.path.join(input_root_dir, emotion))
    file_names = random.sample(file_names, select_per_emotion)
    emotion_2_file_names[emotion] = file_names

os.makedirs(annotation_root_dir, exist_ok=True)

technique_order = []
texts, image_path1_list, image_path2_list = [], [], []
for emotion in emotion_label_list:
    file_names = emotion_2_file_names[emotion]
    for file_name in file_names:
        random.shuffle(output_dirs_and_techniques)
        for output_root_dir, technique in tqdm(output_dirs_and_techniques):
            technique_order.append(technique)

            file_name = file_name.split(".")[0]
            file_path = os.path.join(input_root_dir, emotion, f"{file_name}.JPEG")
            image_path1_list.append(file_path)

            addition_path = os.path.join(addition_root_dir, emotion, file_name, "instruction")
            with open(addition_path) as file:
                addition = "".join(file.readlines())
            texts.append(addition)

            output_path = os.path.join(output_root_dir, emotion, file_name, "0.JPEG")
            image_path2_list.append(output_path)

annotation_path = os.path.join(annotation_root_dir, "annotation.xlsx")
get_annotation(technique_order, texts, image_path1_list, image_path2_list, annotation_path)