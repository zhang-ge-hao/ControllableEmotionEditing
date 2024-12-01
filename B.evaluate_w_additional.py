import os
from tqdm import tqdm
from transformers import AutoModelForImageClassification, AutoFeatureExtractor, pipeline
from PIL import Image
from transformers import CLIPProcessor, CLIPModel
import torch.nn.functional as F
import torch
from openpyxl import Workbook
from openpyxl.styles import Font
from skimage import io, img_as_float, transform
from skimage.metrics import structural_similarity as ssim
import numpy as np
import json
from tqdm import tqdm

import json
import pandas as pd


output_dirs_and_techniques = [
    ("data/5.ultra",    "UltraEdit"),
    ("data/7.brush",    "MagicBrush"),
    ("data/6.ip2p",     "IP2P"),
    ("data/8.2_step",   "2 Steps"),
    ("data/9.no_fsl",   "No FSL"),
    ("data/4.edited",   "Ours"),
]
techniques = [t[1] for t in output_dirs_and_techniques]
emotions = ["amusement", "awe", "contentment", "excitement", "anger", "disgust", "fear", "sadness"]
metrics = ["Delta", "SSIM", "Delta_{SSIM}", "CLIP", "Delta_{CLIP}"]

input_root_dir = "data/0.ilsvrc"

class_count = len(emotions)

def jsonl_to_excel(jsonl_file_path, excel_file_path):
    data = {}
    for metric in metrics:
        data[metric] = []
        data[metric].append([None] + emotions + ["Total"])
        for technique in techniques:
            data[metric].append([technique] + [0] * len(emotions) + [0])

    with open(jsonl_file_path, 'r') as file:
        json_data = {json.loads(line.strip())["Emotion"]: json.loads(line.strip()) for line in file}

    maxs = {metric: [float('-inf')] * len(emotions) + [float('-inf')] for metric in metrics}
    bolds = {metric: [] for metric in metrics}
    for metric in metrics:
        for technique_idx, technique in enumerate(techniques):
            technique_metrics = []
            for emotion_idx, emotion in enumerate(emotions):
                metric_value = json_data[emotion][f"{metric} ({technique})"]
                data[metric][technique_idx + 1][emotion_idx + 1] = metric_value
                technique_metrics.append(metric_value)
                # calculate max
                if round(metric_value, 6) > maxs[metric][emotion_idx]:
                    maxs[metric][emotion_idx] = round(metric_value, 6)
            mean_value = sum(technique_metrics) / len(technique_metrics)
            data[metric][technique_idx + 1][-1] = mean_value
            if round(mean_value, 6) > maxs[metric][-1]:
                maxs[metric][-1] = round(mean_value, 6)
        # bold positions
        for technique_idx in range(len(techniques)):
            for emotion_idx in range(len(emotions) + 1):
                metric_value = data[metric][technique_idx + 1][emotion_idx + 1]
                max_value = maxs[metric][emotion_idx]
                if round(metric_value, 6) == round(max_value, 6):
                    bolds[metric].append((technique_idx, emotion_idx))

    wb = Workbook()
    for metric in metrics:
        ws = wb.create_sheet(title=metric)
        for row in data[metric]:
            ws.append(row)
        for technique_idx, emotion_idx in bolds[metric]:
            cell = ws.cell(row=technique_idx + 2, column=emotion_idx + 2)
            cell.font = Font(bold=True)

    del wb["Sheet"]  # Remove default sheet
    wb.save(excel_file_path)

# EmoSet
classifier_path = "classifier/results/model_1/checkpoint-4726"
model_name = "classifier/results/model_1/checkpoint-4726"
base_model_name = "microsoft/resnet-18"
model = AutoModelForImageClassification.from_pretrained(model_name)
feature_extractor = AutoFeatureExtractor.from_pretrained(base_model_name)
# CLIP
classifier = pipeline("image-classification", model=model, feature_extractor=feature_extractor, device_map="auto", batch_size=1024)
clip_model = CLIPModel.from_pretrained("openai/clip-vit-large-patch14").to("cuda")
clip_processor = CLIPProcessor.from_pretrained("openai/clip-vit-large-patch14")

results = []

for emotion_idx, emotion in tqdm(list(enumerate(emotions))):
    result = {"Emotion": emotion}
    for output_root_dir, title in output_dirs_and_techniques:
        input_file_idxs = []
        output_file_idxs_list = []
        image_files = []
        for file_name in sorted(os.listdir(os.path.join(input_root_dir, emotion))):
            file_name = file_name.split(".")[0]
            input_file_path = os.path.join(input_root_dir, emotion, f"{file_name}.JPEG")
            input_file_idxs.append(len(image_files))
            image_files.append(input_file_path)
            output_file_idxs_list.append([])
            output_dir = os.path.join(output_root_dir, emotion, file_name)
            for output_file_name in sorted(os.listdir(output_dir)):
                output_file_path = os.path.join(output_dir, output_file_name)
                output_file_idxs_list[-1].append(len(image_files))
                image_files.append(output_file_path)
        # EmoSet
        predictions = classifier(image_files, top_k=class_count)
        # CLIP
        images = [Image.open(image_file).convert("RGB") for image_file in image_files]
        inputs = clip_processor(images=images, return_tensors="pt", padding=True).to("cuda")
        with torch.no_grad():
            image_embeddings = clip_model.get_image_features(**inputs)

        scores = []
        for prediction in predictions:
            for d in prediction:
                if d["label"] == f"LABEL_{emotion_idx}":
                    scores.append(d["score"])

        assert len(scores) == image_embeddings.size(0)

        delta_list = []
        clip_list = []
        for input_file_idx, output_file_idxs in zip(input_file_idxs, output_file_idxs_list):
            # delta
            input_score = scores[input_file_idx]
            for output_file_idx in output_file_idxs:
                output_score = scores[output_file_idx]
                delta_list.append(output_score - input_score)
            # CLIP score
            input_embedding = image_embeddings[input_file_idx]
            output_embeddings = image_embeddings[output_file_idxs]
            clip_list.extend(F.cosine_similarity(input_embedding, output_embeddings).tolist())
        
        ssim_list = []
        for file_name in sorted(os.listdir(os.path.join(input_root_dir, emotion))):
            file_name = file_name.split(".")[0]
            input_file_path = os.path.join(input_root_dir, emotion, f"{file_name}.JPEG")
            input_image = io.imread(input_file_path, as_gray=True)
            input_image = transform.resize(input_image, (512, 512), anti_aliasing=True)
            input_image_float = img_as_float(input_image)
            output_dir = os.path.join(output_root_dir, emotion, file_name)
            for output_file_name in sorted(os.listdir(output_dir)):
                output_file_path = os.path.join(output_dir, output_file_name)
                output_image = io.imread(output_file_path, as_gray=True)
                output_image = transform.resize(output_image, (512, 512), anti_aliasing=True)
                output_image_float = img_as_float(output_image)
                ssim_score = ssim(input_image_float, output_image_float, full=True, data_range=1.0)[0]
                ssim_list.append(ssim_score)
        success_delta_list = [d for d in delta_list if d > 0]
        # result[f"Success Rate ({title})"] = sum(success_delta_list) / len(delta_list)
        result[f"Delta ({title})"] = sum(delta_list) / len(delta_list) if len(delta_list) > 0 else 0
        # result[f"Success Delta ({title})"] = sum(success_delta_list) / len(success_delta_list) if len(success_delta_list) > 0 else 0
        result[f"SSIM ({title})"] = sum(ssim_list) / len(ssim_list) if len(ssim_list) > 0 else 0
        result[f"Delta_{{SSIM}} ({title})"] = sum([d * s for d, s in zip(delta_list, ssim_list)]) / len(delta_list) if len(delta_list) > 0 else 0
        result[f"CLIP ({title})"] = sum(clip_list) / len(clip_list) if len(clip_list) > 0 else 0
        result[f"Delta_{{CLIP}} ({title})"] = sum([d * c for d, c in zip(delta_list, clip_list)]) / len(delta_list) if len(delta_list) > 0 else 0
    results.append(result)

with open("data/results.jsonl", "w") as file:
    for result in results:
        file.write(json.dumps(result) + "\n")

jsonl_to_excel("data/results.jsonl", "data/results.xlsx")