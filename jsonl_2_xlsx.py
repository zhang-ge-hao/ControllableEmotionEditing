import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

def jsonl_to_excel(jsonl_file_path, excel_file_path):
    emotions = ["amusement", "awe", "contentment", "excitement", "anger", "disgust", "fear", "sadness"]
    techniques = ["UltraEdit", "MagicBrush", "IP2P", "2 Steps", "Ours"]
    metrics = ["Delta", "SSIM", "Delta_{SSIM}", "CLIP", "Delta_{CLIP}"]
    data = {}
    for metric in metrics:
        data[metric] = []
        data[metric].append([None] + emotions + ["Total"])
        for technique in techniques:
            data[metric].append([technique] + [0] * len(emotions) + [0])

    with open(jsonl_file_path, 'r') as file:
        json_data = {json.loads(line.strip())["Emotion"]: json.loads(line.strip()) for line in file}

    maxs = {metric: [float('-inf')] * len(emotions) + [float('-inf')] for metric in metrics}
    bolds = {metric: [] for metric in metrics}
    for metric in metrics:
        for technique_idx, technique in enumerate(techniques):
            technique_metrics = []
            for emotion_idx, emotion in enumerate(emotions):
                metric_value = json_data[emotion][f"{metric} ({technique})"]
                data[metric][technique_idx + 1][emotion_idx + 1] = metric_value
                technique_metrics.append(metric_value)
                # calculate max
                if round(metric_value, 6) > maxs[metric][emotion_idx]:
                    maxs[metric][emotion_idx] = round(metric_value, 6)
            mean_value = sum(technique_metrics) / len(technique_metrics)
            data[metric][technique_idx + 1][-1] = mean_value
            if round(mean_value, 6) > maxs[metric][-1]:
                maxs[metric][-1] = round(mean_value, 6)
        # bold positions
        for technique_idx in range(len(techniques)):
            for emotion_idx in range(len(emotions) + 1):
                metric_value = data[metric][technique_idx + 1][emotion_idx + 1]
                max_value = maxs[metric][emotion_idx]
                if round(metric_value, 6) == round(max_value, 6):
                    bolds[metric].append((technique_idx, emotion_idx))

    wb = Workbook()
    for metric in metrics:
        ws = wb.create_sheet(title=metric)
        for row in data[metric]:
            ws.append(row)
        for technique_idx, emotion_idx in bolds[metric]:
            cell = ws.cell(row=technique_idx + 2, column=emotion_idx + 2)
            cell.font = Font(bold=True)

    del wb["Sheet"]  # Remove default sheet
    wb.save(excel_file_path)

# Example usage
jsonl_to_excel("data/results.jsonl", "data/results.xlsx")